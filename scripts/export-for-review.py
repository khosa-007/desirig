#!/usr/bin/env python3
"""
Export businesses from Supabase to XLSX for category review.
Groups by category so you can review one category at a time.
Output: data/review/businesses-for-review.xlsx
"""

import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

# Load env
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]

OUTPUT_DIR = Path(__file__).parent.parent / "data" / "review"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_FILE = OUTPUT_DIR / "businesses-for-review.xlsx"


def fetch_all(supabase, table, select="*", batch_size=1000):
    """Fetch all rows from a table using pagination."""
    all_rows = []
    offset = 0
    while True:
        resp = supabase.table(table).select(select).range(offset, offset + batch_size - 1).execute()
        rows = resp.data
        if not rows:
            break
        all_rows.extend(rows)
        offset += batch_size
        if len(rows) < batch_size:
            break
    return all_rows


def main():
    print("Connecting to Supabase...")
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Fetch categories
    print("Fetching categories...")
    cats = supabase.table("categories").select("id, name, slug, is_trucking, parent_id").execute().data
    cat_map = {c["id"]: c for c in cats}

    # Fetch cities
    print("Fetching cities...")
    cities = supabase.table("cities").select("id, name, province").execute().data
    city_map = {c["id"]: c for c in cities}

    # Fetch all businesses
    print("Fetching businesses (this may take a moment)...")
    businesses = fetch_all(supabase, "businesses",
        "id, name, address, phone, website, google_rating, google_review_count, category_id, city_id, province, is_desi_owned, business_status")

    print(f"  Found {len(businesses)} businesses")

    # Group by category
    by_category = defaultdict(list)
    for b in businesses:
        cat = cat_map.get(b["category_id"], {})
        cat_name = cat.get("name", "Unknown")
        by_category[cat_name].append(b)

    # Create workbook
    print("Creating XLSX...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    header_fill = PatternFill(start_color="FACC15", end_color="FACC15", fill_type="solid")
    header_font = Font(bold=True, size=11)

    headers = [
        "Name", "Address", "City", "Province", "Phone", "Website",
        "Google Rating", "Reviews", "Current Category", "Is Trucking",
        "Verified Category", "Sub Category", "Notes", "Keep"
    ]

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Category", "Count", "Is Trucking"])
    for cat_name in sorted(by_category.keys()):
        items = by_category[cat_name]
        cat_info = next((c for c in cats if c["name"] == cat_name), {})
        ws_summary.append([cat_name, len(items), "Yes" if cat_info.get("is_trucking") else "No"])

    # Style summary header
    for col in range(1, 4):
        cell = ws_summary.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # One sheet per category
    for cat_name in sorted(by_category.keys()):
        items = by_category[cat_name]
        cat_info = next((c for c in cats if c["name"] == cat_name), {})

        # Sheet names max 31 chars
        sheet_name = cat_name[:31].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(sheet_name)

        # Header row
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)

        # Data rows
        for b in sorted(items, key=lambda x: x.get("name", "")):
            city_info = city_map.get(b.get("city_id"), {})
            ws.append([
                b.get("name", ""),
                b.get("address", ""),
                city_info.get("name", ""),
                b.get("province", ""),
                b.get("phone", ""),
                b.get("website", ""),
                b.get("google_rating"),
                b.get("google_review_count", 0),
                cat_name,
                "Yes" if cat_info.get("is_trucking") else "No",
                "",  # verified_category (empty for review)
                "",  # sub_category (empty for review)
                "",  # notes (empty for review)
                "YES",  # keep (default YES)
            ])

        # Auto-width (approximate)
        for col_idx, _ in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    wb.save(str(OUTPUT_FILE))
    print(f"\nExported to: {OUTPUT_FILE}")
    print(f"Total businesses: {len(businesses)}")
    print(f"Categories: {len(by_category)}")
    print(f"Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
